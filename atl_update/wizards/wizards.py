# -*- coding: utf-8 -*-
#############################################
###############  #######################
############################################

from openerp import models, fields, api, _
from openpyxl import load_workbook
import os 
import urllib
# import subprocess

class atl_updation_wizard(models.TransientModel):
    _name = "atl.updation.wizard"
    
    @api.multi
    def create_request(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        if "%s/fbr_data/test.xlsx" % dir_path[:-8]:
            os.system("rm -r %s/fbr_data/test.xlsx" % dir_path[:-8])
        # dls = "http://download1.fbr.gov.pk/Docs/20171169130811820170115-IT-ATL-2015.xlsx"
        dls = "http://download1.fbr.gov.pk/Docs/2017123101393794320170122-IT-ATL-2015.xlsx"
        urllib.urlretrieve(dls, "%s/fbr_data/test.xlsx" % dir_path[:-8])
        print 'Page Download successfully'

        cc = self.return_column_from_excel('%s/fbr_data/test.xlsx' % dir_path[:-8], 2, 1)

        active_class =self.env['res.partner'].search([])
        for x in active_class:
            if x.registration_no in cc:
                x.filer = 1

    def return_column_from_excel(self,file_name, column_num, first_data_row):
        registration , registration_no = ([],[])
        wb = load_workbook(filename=file_name)
        print 'Page load successfully'
        no_of_files = len(wb.get_sheet_names())
        no = 1
        while no < no_of_files + 1:
            ws = wb.get_sheet_by_name('Part-%s' % no)
            min_col, min_row, max_col, max_row = (column_num, first_data_row, column_num, ws.max_row)
            part = ws.get_squared_range(min_col, min_row, max_col, max_row)
            itr = 0
            for x in part:
                if itr > 1:
                    registration.append(x)
                itr = itr + 1
            no = no + 1
        print 'Itration with pages completed'
        
        for rec in registration:
            for x in rec:
                registration_no.append(str(x.value))
        return registration_no